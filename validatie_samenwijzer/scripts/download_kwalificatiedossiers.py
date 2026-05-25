"""Download kwalificatiedossier-PDFs uit s-bb.nl zips, gemapt op crebo.

Vereist dat de 4 alfabetische zips en de crebolijsten al gedownload zijn in
``kwalificatiedossiers/``. Resultaat: ``kwalificatiedossiers/pdfs/<crebo>.pdf``.
"""

from __future__ import annotations

import json
import re
import sqlite3
import sys
import unicodedata
import zipfile
from difflib import get_close_matches
from pathlib import Path

import openpyxl
import pdfplumber

ROOT = Path(__file__).resolve().parents[2]
KWAL_DIR = ROOT / "kwalificatiedossiers"
PDF_DIR = KWAL_DIR / "pdfs"
DB_PATH = ROOT / "validatie_samenwijzer" / "data" / "validatie.db"

# Crebo's die niet in de s-bb crebolijsten staan maar wel via s-bb-search
# herleidbaar zijn naar een bestaande dossier-PDF (handmatig vastgesteld).
HANDMATIGE_OVERRIDES: dict[str, str] = {
    "25898": "Engineering",
    "25923": "Meubels en (scheeps)interieurs maken",
    "25924": "Meubels en (scheeps)interieurs maken",
    "25926": "Meubels en (scheeps)interieurs maken",
    "25928": "Meubels en (scheeps)interieurs maken",
    "25958": "Sociaal werk",
    "25960": "Dienstverlening",
    "25961": "Dienstverlening",
    "25977": "Voeding",
    "25981": "Agro productie, handel en technologie",
    "25983": "Agro productie, handel en technologie",
    "25987": "Agro productie, handel en technologie",
    "25992": "Groen, grond en groene infra",
    "25997": "Ondernemerschap op basis van een vakspecialisme",
    "25998": "Software development",
    "25999": "ICT support",
    "27002": "Signmaking",
    "27004": "Signmaking",
}


def normaliseer(naam: str) -> str:
    """Vergelijk-vriendelijke normalisatie van een dossiernaam."""

    nfkd = unicodedata.normalize("NFKD", naam)
    zonder_accent = "".join(c for c in nfkd if not unicodedata.combining(c))
    # mojibake fix: in filenames is 'ë' soms 'δ'
    zonder_accent = zonder_accent.replace("δ", "e")
    zonder_accent = re.sub(r"[/\\\-]+", "", zonder_accent)
    zonder_accent = re.sub(r"\s+", "", zonder_accent)
    return zonder_accent.lower()


def laad_crebos_uit_db() -> set[str]:
    conn = sqlite3.connect(DB_PATH)
    try:
        rows = conn.execute(
            "SELECT DISTINCT crebo FROM oer_documenten WHERE crebo IS NOT NULL"
        ).fetchall()
    finally:
        conn.close()
    return {r[0] for r in rows}


def laad_crebo_mapping() -> dict[str, str]:
    """Map kwalificatie-crebo → dossiernaam uit alle gedownloade crebolijsten."""

    mapping: dict[str, tuple[str, str]] = {}  # crebo -> (jaar, dossiernaam)
    vervangingen: dict[str, str] = {}  # oude_crebo -> nieuwe_crebo
    for xlsx in sorted((KWAL_DIR / "lijsten").glob("crebo_*.xlsx")):
        jaar = xlsx.stem.split("_", 1)[1]
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        if "Complete lijst" in wb.sheetnames:
            ws = wb["Complete lijst"]
            huidige_dossier: str | None = None
            for row in ws.iter_rows(values_only=True):
                dossier_crebo = row[0]
                dossiernaam = row[2]
                kwal_crebo = row[3]
                if isinstance(dossiernaam, str):
                    huidige_dossier = dossiernaam.strip()
                for crebo_val in (dossier_crebo, kwal_crebo):
                    if isinstance(crebo_val, int) and huidige_dossier:
                        key = str(crebo_val)
                        bestaand = mapping.get(key)
                        if bestaand is None or jaar > bestaand[0]:
                            mapping[key] = (jaar, huidige_dossier)
        # Vervallen sheet: oude crebo (kol A of F) → vervangen door (kol F)
        for sheet in ("Vervallen", "Wijzigingen"):
            if sheet not in wb.sheetnames:
                continue
            for row in wb[sheet].iter_rows(values_only=True):
                if not isinstance(row[0], int):
                    continue
                oud = str(row[0])
                nieuw = row[5] if len(row) > 5 else None
                dossier = row[1] if isinstance(row[1], str) else None
                if isinstance(nieuw, int):
                    vervangingen[oud] = str(nieuw)
                if dossier and oud not in mapping:
                    mapping[oud] = (jaar, dossier.strip())

    resultaat = {k: v[1] for k, v in mapping.items()}
    # transitive close: als A → B en B in resultaat, dan ook A → naam van B
    for oud, nieuw in vervangingen.items():
        if oud not in resultaat and nieuw in resultaat:
            resultaat[oud] = resultaat[nieuw]
    resultaat.update(HANDMATIGE_OVERRIDES)
    return resultaat


def bouw_zip_index() -> dict[str, list[tuple[Path, str]]]:
    """Map genormaliseerde stem → lijst van (zip-pad, originele filename)."""

    idx: dict[str, list[tuple[Path, str]]] = {}
    for zp in sorted(KWAL_DIR.glob("*.zip")):
        with zipfile.ZipFile(zp) as zf:
            for naam in zf.namelist():
                if not naam.lower().endswith(".pdf"):
                    continue
                stem_norm = normaliseer(Path(naam).stem)
                idx.setdefault(stem_norm, []).append((zp, naam))
    return idx


def vind_pdfs_voor_dossier(
    dossiernaam: str, zip_index: dict[str, list[tuple[Path, str]]]
) -> list[tuple[Path, str]]:
    sleutel = normaliseer(dossiernaam)
    if sleutel in zip_index:
        return zip_index[sleutel]
    matches = get_close_matches(sleutel, zip_index.keys(), n=1, cutoff=0.85)
    if matches:
        return zip_index[matches[0]]
    return []


_DATUM_RE = re.compile(r"Geldig vanaf\s+(\d{2}-\d{2}-\d{4})")


def extract_geldig_vanaf(pdf_bytes: bytes, tmp: Path) -> str:
    """Lees 'Geldig vanaf dd-mm-jjjj' van pagina 1; lege string bij falen."""

    tmp.write_bytes(pdf_bytes)
    try:
        with pdfplumber.open(tmp) as pdf:
            tekst = pdf.pages[0].extract_text() or ""
    except Exception:
        return ""
    match = _DATUM_RE.search(tekst)
    if not match:
        return ""
    dag, maand, jaar = match.group(1).split("-")
    return f"{jaar}-{maand}-{dag}"


def kies_meest_recente(
    kandidaten: list[tuple[Path, str]], tmp: Path
) -> tuple[Path, str, str]:
    """Geef (zip, filename, geldig_vanaf) voor meest recente PDF."""

    if len(kandidaten) == 1:
        zp, fn = kandidaten[0]
        with zipfile.ZipFile(zp) as zf:
            geldig = extract_geldig_vanaf(zf.read(fn), tmp)
        return zp, fn, geldig

    beste: tuple[Path, str, str] | None = None
    for zp, fn in kandidaten:
        with zipfile.ZipFile(zp) as zf:
            geldig = extract_geldig_vanaf(zf.read(fn), tmp)
        if beste is None or geldig > beste[2]:
            beste = (zp, fn, geldig)
    assert beste is not None
    return beste


def main() -> int:
    PDF_DIR.mkdir(parents=True, exist_ok=True)

    onze_crebos = laad_crebos_uit_db()
    crebo_naar_dossier = laad_crebo_mapping()
    zip_index = bouw_zip_index()

    matched = onze_crebos & crebo_naar_dossier.keys()
    excel_unmatched = sorted(onze_crebos - crebo_naar_dossier.keys())

    tmp_pdf = KWAL_DIR / "_tmp.pdf"
    resultaten: dict[str, dict] = {}
    geen_pdf: list[tuple[str, str]] = []

    for i, crebo in enumerate(sorted(matched), 1):
        dossier = crebo_naar_dossier[crebo]
        kandidaten = vind_pdfs_voor_dossier(dossier, zip_index)
        if not kandidaten:
            geen_pdf.append((crebo, dossier))
            continue
        zp, fn, geldig = kies_meest_recente(kandidaten, tmp_pdf)
        doel = PDF_DIR / f"{crebo}.pdf"
        with zipfile.ZipFile(zp) as zf:
            doel.write_bytes(zf.read(fn))
        resultaten[crebo] = {
            "dossiernaam": dossier,
            "bron_zip": zp.name,
            "bron_pdf": fn,
            "geldig_vanaf": geldig,
            "doel": doel.name,
        }
        if i % 25 == 0:
            print(f"  {i}/{len(matched)} crebos verwerkt")

    if tmp_pdf.exists():
        tmp_pdf.unlink()

    rapport = {
        "totaal_crebos_db": len(onze_crebos),
        "gemapt_via_excel": len(matched),
        "opgeslagen_pdfs": len(resultaten),
        "geen_pdf_match": geen_pdf,
        "excel_unmatched": excel_unmatched,
        "resultaten": resultaten,
    }
    (KWAL_DIR / "download_rapport.json").write_text(
        json.dumps(rapport, indent=2, ensure_ascii=False)
    )
    print(
        f"\nKlaar: {len(resultaten)} PDFs opgeslagen in {PDF_DIR}. "
        f"Geen PDF voor {len(geen_pdf)} crebos, "
        f"{len(excel_unmatched)} crebos niet in crebolijsten."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
